from tkinter import Tk, filedialog, Label, Button
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows 
from openpyxl.styles import Font, PatternFill, Border, Side
import os
import pandas as pd
import aspose.zip as az
import re

mensagem = "Aguardando ações..."

#----- INTERFACE GRÁFICA -----

#Função da interface gráfica.
def create_gui():

    global label_mensagem

    #Define root como inicializador da interface.
    root = Tk()
    root.title("Migração AdvBox")

    Label(root, text="Primeiro. clique em: Upload dos arquivos .rar.").pack()
    Label(root, text="Após a mensagem de sucesso, clique em: Processar arquivos.").pack()
    Label(root, text="Os arquivos processados estarao em dados_migracao.").pack()
    Label(root, text="Obs. Espere o status Processamento concluido").pack()
    Label(root, text=" ").pack()
    
    label_mensagem = Label(root, text=f"Status: {mensagem}")
    label_mensagem.pack()
    Label(root, text=" ").pack()

    Label(root, text="Selecione o arquivo de backup: ").pack()
    Button(root, text="Upload dos arquivos", command=upload_file).pack()
    Button(root, text="Processar arquivos", command=process_file).pack()

    #A janela não fechará após executar as ações.
    root.mainloop()

#----- FUNÇÕES DOS BOTÕES -----

#Função para encontrar o caminho do arquivo rar.
def upload_file():
    
    global mensagem
    global backup_path

    # Procure por arquivo .rar.
    try:
        backup_path = filedialog.askopenfilename(filetype=[("RAR files", "*.rar")])
        
        # Se arquivo backup encontrado.
        if backup_path:
            
            # Feedback
            print("Arquivo carregado com sucesso")
            mensagem = "Arquivo carregado com sucesso"
            atualizar_mensagem()

        # Se arquivo backup não-encontrado.
        else:

            # Feedback
            print("Arquivo não encontrado")
            mensagem = "Arquivo não encontrado"
            atualizar_mensagem()

    except:
        print("Erro ao carregar o arquivo")
        mensagem = "Erro ao carregar o arquivo"
        atualizar_mensagem()

#Função de realizar o processo de transformação dos dados do backup.
def process_file():

    global mensagem
    
    if backup_path:

        #------ Processo de extração e transformação dos arquivos -----
        extract_path = "./backup" # Define o nome do caminho atual + nome da pasta.
        output_path = "./dados_migracao"
        os.makedirs(output_path, exist_ok=True) # Cria um novo diretório no caminho determinado.

        # Extração do arquivo rar
        try:
            
            extract_rar(backup_path, extract_path)
            print("Arquivo rar extraído com sucesso!")
            mensagem = "Arquivo rar extraído com sucesso!"
            atualizar_mensagem()
        
        except Exception as e:
            
            print(f"Erro ao extrair arquivo{e}")
            mensagem = f"Erro ao extrair arquivo{e}"
            atualizar_mensagem()
        
        # Processamento dos dados CLIENTES
        try:
            
            transform_dataClientes(extract_path, output_path) 
            print("Dados clientes processados!")
            mensagem = "Dados clientes processado!"
            atualizar_mensagem()
        
        except Exception as e:
            
            print(f"Erro ao processar dados clientes{e}")
            mensagem = f"Erro ao processar dados clientes{e}"
            atualizar_mensagem()

        # Processamento dos dados PROCESSOS
        try:

            transform_dataProcess(extract_path, output_path) 
            print("Dados processos processados!")
            mensagem = "Dados processos processado!"
            atualizar_mensagem()
        
        except Exception as e:
            
            print(f"Erro ao processar dados processos{e}")
            mensagem = f"Erro ao processar dados processos{e}"
            atualizar_mensagem()
        
        try:

            # Leitura dos arquivos csv.
            clientes_csv = pd.read_csv(os.path.join(output_path, 'CLIENTES.csv'))
            processos_csv = pd.read_csv(os.path.join(output_path, 'PROCESSOS.csv'))

            # Convertendo para excel.
            clientes_csv.to_excel(os.path.join(output_path, 'CLIENTES.xlsx'), index = False)
            processos_csv.to_excel(os.path.join(output_path, 'PROCESSOS.xlsx'), index = False) 

            # Fazendo leitura excel.
            clientes_excel = pd.read_excel(os.path.join(output_path, 'CLIENTES.xlsx'))
            processos_excel = pd.read_excel(os.path.join(output_path, 'PROCESSOS.xlsx'))

            # Manipulação excel.
            wb = Workbook() # Cria nova página.
            ws = wb.active # Seleciona a página ativa.
            ws.title = "MIGRACAO_NOVO" # Define o título da página.

            # Definindo fonte negrito.
            negrito = Font(bold=True)
            # Definindo cinza claro como background.
            cor_fundo = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid") 
            # Definir borda para o cabeçalho
            borda = Border(
                top=Side(style="thin"),
                bottom=Side(style="thin"),
                left=Side(style="thin"),
                right=Side(style="thin")
            )

            # Adiciona o título da planilha clientes.
            ws.append(["PLANILHA CLIENTES"])
            cell = ws.cell(row=1, column=1)
            cell.font = negrito
            # Adiciona todos os dados clientes.
            for linha in dataframe_to_rows(clientes_excel, index = False, header = True):
                ws.append(linha)
            # Pinta o cabeçalho da planilha CLIENTES
            for col_num, cell in enumerate(ws[2], 1):  # O cabeçalho começa na linha 2
                cell.font = negrito  # Aplica o negrito
                cell.fill = cor_fundo  # Aplica a cor de fundo
                cell.border = borda  # Aplica a borda

            # Adiciona um espaço entre as planilhas
            ws.append([])

            # Adiciona o título da planilha processos.
            ws.append(["PLANILHAS PROCESSOS"])
            cell = ws.cell(row=ws.max_row, column=ws.max_column)
            cell.font = negrito
            # Adiciona todos os dados processos.
            for linha in dataframe_to_rows(processos_excel, index = False, header = True):
                ws.append(linha)
            # Pinta o cabeçalho da planilha PROCESSOS
            for col_num, cell in enumerate(ws[ws.max_row + 1], 1):  # O cabeçalho começa na linha depois da última linha
                cell.font = negrito  # Aplica o negrito
                cell.fill = cor_fundo  # Aplica a cor de fundo
                cell.border = borda  # Aplica a borda
                
            
            # Salva a nova planilha migracao.
            wb.save(os.path.join(output_path, 'MIGRACAO_NOVO.xlsx'))   

            print("Planilhas geradas")
            mensagem = "Planilhas geradas"
            atualizar_mensagem()
        
        except Exception as e:
            
            print(f"Erro ao processar planilhas{e}")
            mensagem = f"Erro ao processar planilhas{e}"
            atualizar_mensagem()

        print("Processamento concluido")
        mensagem = "Processamento concluido"
        atualizar_mensagem()

    else:

        print("Nao ha arquivos a serem processado")
        mensagem = "Nao ha arquivos a serem processado"
        atualizar_mensagem(mensagem)

#Atualiza a label_mensagem que exibe os status do processo.
def atualizar_mensagem():
    
    label_mensagem.config(text=f"Status: {mensagem}")

#----- Funções para process_file -----

# Extração arquivo .rar
def extract_rar(backup_path, extract_path):

    os.makedirs(extract_path, exist_ok=True) # Cria um novo diretório no caminho determinado.

    # Extrcação do conteúdo .rar
    with az.rar.RarArchive(backup_path) as archive:
        archive.extract_to_directory(extract_path)
    
    # Listar os arquivos extraídos
    extracted_files = os.listdir(extract_path)
    print(extracted_files)
    atualizar_mensagem()

# Processamento dos arquivos advbox.

def transform_dataClientes(extract_path, output_path):

    #----- Processamento dados clientes -----

    # Leitura dos dados gerais dos clientes.
    clientes_CodEmpresa = pd.read_csv(os.path.join(extract_path, 'v_clientes_CodEmpresa_92577.csv'), encoding="latin1", sep=";")

    # Filtro somente colunas necessarias dados clientes.
    clientes_CodEmpresa_formatado = clientes_CodEmpresa[["razao_social","razao_social_2","contato_nome", "cpf_cnpj", "rg", "nacionalidade", 
                                                     "nascimento", "estado_civil", "profissao", "contato_telefone1", "telefone1", "telefone2",
                                                     "telefone3","contato_telefone2", "telefone_comercial", "email1", "email2", "uf", "cidade", 
                                                     "bairro", "cep", "pis", "nome_mae", "observacoes"]]

    # Convertendo coluna nascimento para formato datetime no modelo DD/MM/AAAA
    clientes_CodEmpresa_formatado["nascimento"] = pd.to_datetime(clientes_CodEmpresa_formatado["nascimento"]).dt.strftime("%d/%m/%Y")

    # Removendo caracteres do campo rg.
    clientes_CodEmpresa_formatado["rg"] = clientes_CodEmpresa_formatado["rg"].replace("UF", "", regex=True)

    # Verifica se a nacionalidade termina em ro ou ra, para assim, aferir seu sexo.
    clientes_CodEmpresa_formatado["sexo"] = "Sem_dados"

    # Armazena a função regex na variavéis correspondentes.
    masculino = re.compile(r'[A-Za-z]+ro') 
    feminino = re.compile(r'[A-Za-z]+ra')

    # Percorre todos os índices e dados da coluna nacionalidade.
    for i, nacionalidade in enumerate(clientes_CodEmpresa_formatado["nacionalidade"]): 
        
        # Se o dado for string.
        if isinstance(nacionalidade, str):
            # Se identificado ro.
            if masculino.search(nacionalidade):
                clientes_CodEmpresa_formatado.at[i, "sexo"] = "M" # Coloque "M" no índice i da coluna sexo.
            # Se identificado ra.
            elif feminino.search(nacionalidade):
                clientes_CodEmpresa_formatado.at[i, "sexo"] = "F" # Coloque "F" no índice i da coluna sexo.
    
    #Código do estado civil dos clientes.
    clientes_estadoCivil = pd.read_csv(os.path.join(extract_path,"v_cliente_estado_civil_CodEmpresa_92577.csv"), 
                                   encoding = "latin1", sep=";")

    # Left Join estado civil
    clientes_CodEmpresa_formatado1 = pd.merge(clientes_CodEmpresa_formatado, clientes_estadoCivil[["sigla", "descricao"]], 
                                                        left_on = "estado_civil", right_on = "sigla", how = "left")

    clientes_CodEmpresa_formatado1 = clientes_CodEmpresa_formatado1.drop(axis=1, columns="estado_civil")
    clientes_CodEmpresa_formatado1 = clientes_CodEmpresa_formatado1.drop(axis=1, columns="sigla")

    clientes_CodEmpresa_formatado1 = clientes_CodEmpresa_formatado1.rename(columns={"descricao": "estado_civil"})

    clientes_CodEmpresa_formatado1.to_csv(os.path.join(output_path, 'CLIENTES.csv')) 

def transform_dataProcess(extract_path, output_path):

    #----- Processamento dados processos -----

    # Carregamento dados gerais dos processos.
    processos_CodEmpresa = pd.read_csv(os.path.join(extract_path, 'v_processos_CodEmpresa_92577.csv'), 
            encoding="latin1", sep=";")

    processos_CodEmpresa_filtrado = processos_CodEmpresa[["cod_cliente", "codigo", "tipo","tipoprocesso","grupo_processo", "codigo_fase", 
                                                      "statusprocessual", "numero_processo", "codorigem","numero_vara", "codcomarca", 
                                                      "valor_causa","valor_causa2", "pasta", "data_contratacao", "data_distribuicao", 
                                                      "data_encerramento", "data_ultima_visualizacao", "observacoes"
                                                      ]] 
    
    # Leitura dos dados gerais dos clientes.
    clientes_CodEmpresa = pd.read_csv(os.path.join(extract_path, 'v_clientes_CodEmpresa_92577.csv'), encoding="latin1", sep=";")

    # Left Join para nomes dos clientes na planilha processos.
    processos_CodEmpresa_filtrado_nome = pd.merge(processos_CodEmpresa_filtrado, clientes_CodEmpresa[["codigo", "razao_social"]], left_on = "cod_cliente", right_on = "codigo", how = "left")

    # Exclui dados dos códigos.
    processos_CodEmpresa_filtrado_nome = processos_CodEmpresa_filtrado_nome.drop(columns=["codigo_x"])
    processos_CodEmpresa_filtrado_nome = processos_CodEmpresa_filtrado_nome.drop(columns=["codigo_y"])

    # Transformando colunas datas em datetime no modelo DD/MM/AAAA
    processos_CodEmpresa_filtrado_nome["data_contratacao"] = pd.to_datetime(processos_CodEmpresa_filtrado_nome["data_contratacao"]).dt.strftime("%d/%m/%Y")
    
    processos_CodEmpresa_filtrado_nome["data_distribuicao"] = pd.to_datetime(processos_CodEmpresa_filtrado_nome["data_distribuicao"]).dt.strftime("%d/%m/%Y")

    processos_CodEmpresa_filtrado_nome["data_ultima_visualizacao"] = pd.to_datetime(processos_CodEmpresa_filtrado_nome["data_ultima_visualizacao"]).dt.strftime("%d/%m/%Y")

    # Adiciona dados dos statusprocessuais
    statusprocessual = pd.read_csv(os.path.join(extract_path, "v_statusprocessual_CodEmpresa_92577.csv"), 
                                  encoding = "latin1", sep=";")

    # Left Join status processual
    processos_CodEmpresa_filtrado_nome_status = pd.merge(processos_CodEmpresa_filtrado_nome, statusprocessual[["codigo", "descricao"]], 
                                                     left_on = "statusprocessual", right_on = "codigo", how = "left")

    # Excluindo colunas com código
    processos_CodEmpresa_filtrado_nome_status = processos_CodEmpresa_filtrado_nome_status.drop(columns="statusprocessual")
    processos_CodEmpresa_filtrado_nome_status = processos_CodEmpresa_filtrado_nome_status.drop(columns="codigo")

    # Renomeando nome da tabela para o nome presente na planilha modelo.
    processos_CodEmpresa_filtrado_nome_status = processos_CodEmpresa_filtrado_nome_status.rename(columns={"descricao": "FASE_PROCESSUAL"})

    #Dados da comarca.
    comarca = pd.read_csv(os.path.join(extract_path, "v_comarca_CodEmpresa_92577.csv"), 
                                  encoding = "latin1", sep=";")
    
    # Left Join comarca
    processos_CodEmpresa_filtrado_nome_status_comarca = pd.merge(processos_CodEmpresa_filtrado_nome_status, comarca[["codigo", "descricao"]], 
                                                     left_on = "codcomarca", right_on = "codigo", how = "left")
    # Excluindo colunas código
    processos_CodEmpresa_filtrado_nome_status_comarca = processos_CodEmpresa_filtrado_nome_status_comarca.drop(columns="codcomarca")
    processos_CodEmpresa_filtrado_nome_status_comarca = processos_CodEmpresa_filtrado_nome_status_comarca.drop(columns="codigo")

    # Renomeando coluna de acordo com o modelo proposto
    processos_CodEmpresa_filtrado_nome_status_comarca = processos_CodEmpresa_filtrado_nome_status_comarca.rename(columns={"descricao": "COMARCA"})

    #Dados dos grupo processo.
    grupo_processo = pd.read_csv(os.path.join(extract_path, "v_grupo_processo_CodEmpresa_92577.csv"), 
                                  encoding = "latin1", sep=";")
    # Left Join grupo processo.
    processos_CodEmpresa_filtrado_nsc_gp = pd.merge(processos_CodEmpresa_filtrado_nome_status_comarca, grupo_processo[["codigo", "descricao"]], 
                                                     left_on = "grupo_processo", right_on = "codigo", how = "left")
    # Excluindo colunas código.
    processos_CodEmpresa_filtrado_nsc_gp = processos_CodEmpresa_filtrado_nsc_gp.drop(columns="grupo_processo")
    processos_CodEmpresa_filtrado_nsc_gp = processos_CodEmpresa_filtrado_nsc_gp.drop(columns="codigo")

    # Renomeando coluna de acordo com o modelo
    processos_CodEmpresa_filtrado_nsc_gp = processos_CodEmpresa_filtrado_nsc_gp.rename(columns={"descricao": "GRUPO_ACAO"})

    # Preenche todos os dados 0 com administrativo (Conforme pedido pelo cliente)
    processos_CodEmpresa_filtrado_nsc_gp["GRUPO_ACAO"] = processos_CodEmpresa_filtrado_nsc_gp["GRUPO_ACAO"].fillna("Administrativo")

    #Dados da fase processual.
    fase = pd.read_csv(os.path.join(extract_path, "v_fase_CodEmpresa_92577.csv"), 
                                  encoding = "latin1", sep=";")
    
    # Left Join fase processual.
    processos_CodEmpresa_filtrado_nsc_gpfase = pd.merge(processos_CodEmpresa_filtrado_nsc_gp, fase[["codigo", "fase"]], 
                                                     left_on = "codigo_fase", right_on = "codigo", how = "left")

    # Excluir colunas com código.
    processos_CodEmpresa_filtrado_nsc_gpfase = processos_CodEmpresa_filtrado_nsc_gpfase.drop(columns="codigo_fase")
    processos_CodEmpresa_filtrado_nsc_gpfase = processos_CodEmpresa_filtrado_nsc_gpfase.drop(columns="codigo")

    # Renomear coluna conforme pedido.
    processos_CodEmpresa_filtrado_nsc_gpfase = processos_CodEmpresa_filtrado_nsc_gpfase.rename(columns={"fase": "GRUPO_FASE"})

    # Salvando arquivo processos.
    processos_CodEmpresa_filtrado_nsc_gpfase.to_csv(os.path.join(output_path, 'PROCESSOS.csv'))

if __name__ == '__main__':
    create_gui()

